#!/usr/bin/env python3
"""
==============================================================================
RCBT - Product Excel Export
==============================================================================

Input
-----
analysis/product/product_list.json
analysis/product/product_metadata_normalized.json
analysis/vendor/vendor_mapping.json

Output
------
output/Product_List.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter



# ==============================================================================
# PATH
# ==============================================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

PRODUCT_LIST_FILE = (
    PROJECT_ROOT
    / "analysis/product/product_list.json"
)

PRODUCT_METADATA_FILE = (
    PROJECT_ROOT
    / "analysis/product/product_metadata_normalized.json"
)

VENDOR_MAPPING_FILE = (
    PROJECT_ROOT
    / "analysis/vendor/vendor_mapping.json"
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)

OUTPUT_FILE = (
    OUTPUT_DIR
    / "Product_List.xlsx"
)



# ==============================================================================
# EXCEL HEADER
# ==============================================================================

HEADERS = [

    "No",

    "Brand",

    "Official Vendor",

    "Official Site",

    "Product Type",

    "Product Model",

    "Normalized Product",

    "Description",

    "Quantity",

    "Unit Price",

    "Total Price",

    "Remarks",

    "Photo",
]



# ==============================================================================
# STYLE
# ==============================================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

BODY_FONT = Font(
    name="Calibri",
    size=11,
)

THIN_BORDER = Border(

    left=Side(style="thin"),

    right=Side(style="thin"),

    top=Side(style="thin"),

    bottom=Side(style="thin"),
)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

TOP = Alignment(
    vertical="top",
)

WRAP = Alignment(
    vertical="top",
    wrap_text=True,
)



# ==============================================================================
# JSON
# ==============================================================================

def load_json(path: Path):

    if not path.exists():

        raise FileNotFoundError(path)

    with open(
        path,
        "r",
        encoding="utf-8",
    ) as f:

        return json.load(f)



# ==============================================================================
# LOOKUP
# ==============================================================================

def build_vendor_lookup(mapping):

    """
    vendor_mapping.json

    {
        "Mikrotik": {
            "official_name": "...",
            "official_site": "..."
        }
    }
    """

    return mapping



def build_product_lookup(products):

    """
    Lookup berdasarkan:
        (brand, product_model)

    Menghindari collision jika ada model sama
    dari vendor berbeda.
    """

    lookup = {}

    for item in products:

        key = (

            str(
                item.get(
                    "brand",
                    "",
                )
            ).strip().lower(),

            str(
                item.get(
                    "product_model",
                    "",
                )
            ).strip().lower(),

        )

        lookup[key] = item

    return lookup

# ==============================================================================
# HEADER
# ==============================================================================

def write_header(ws):

    for column, header in enumerate(
        HEADERS,
        start=1,
    ):

        cell = ws.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER



# ==============================================================================
# PRODUCT
# ==============================================================================

def write_products(
    ws,
    products,
    vendor_lookup,
    product_lookup,
):

    row = 2

    for number, product in enumerate(
        products,
        start=1,
    ):

        brand = str(
            product.get(
                "Brand",
                "",
            )
        ).strip()

        model = str(
            product.get(
                "Product Model",
                "",
            )
        ).strip()

        vendor = vendor_lookup.get(
            brand,
            {},
        )

        lookup_key = (
            brand.lower(),
            model.lower(),
        )

        metadata = product_lookup.get(
            lookup_key,
            {},
        )

        official_vendor = vendor.get(
            "official_name",
            brand,
        )

        official_site = vendor.get(
            "official_site",
            "",
        )

        normalized_product = metadata.get(
            "normalized_product",
            model,
        )

        values = [

            number,

            brand,

            official_vendor,

            official_site,

            product.get(
                "Product Type",
                "",
            ),

            model,

            normalized_product,

            product.get(
                "Description",
                "",
            ),

            product.get(
                "Quantity/Unit",
                "",
            ),

            product.get(
                "Unit Price",
                "",
            ),

            product.get(
                "Total Price",
                "",
            ),

            product.get(
                "Remarks",
                "",
            ),

            "Open Photo",
        ]

        for column, value in enumerate(
            values,
            start=1,
        ):

            cell = ws.cell(
                row=row,
                column=column,
                value=value,
            )

            cell.font = BODY_FONT
            cell.border = THIN_BORDER

            if column in (
                1,
                2,
                3,
                5,
                6,
                9,
            ):
                cell.alignment = CENTER

            elif column == 8:
                cell.alignment = WRAP

            else:
                cell.alignment = TOP

        # ----------------------------------------------------------
        # Official Site Hyperlink
        # ----------------------------------------------------------

        if official_site:

            site_cell = ws.cell(
                row=row,
                column=4,
            )

            site_cell.hyperlink = official_site
            site_cell.style = "Hyperlink"
            site_cell.value = official_site

        # ----------------------------------------------------------
        # Product Photo Hyperlink
        # ----------------------------------------------------------

        photo = product.get(
            "Product Photo",
            "",
        )

        if photo:

            photo_cell = ws.cell(
                row=row,
                column=13,
            )

            photo_cell.hyperlink = photo
            photo_cell.style = "Hyperlink"
            photo_cell.value = "Open Photo"

        row += 1

# ==============================================================================
# WORKSHEET FORMAT
# ==============================================================================

def auto_width(ws):

    for column_cells in ws.columns:

        column_letter = get_column_letter(
            column_cells[0].column
        )

        max_length = 0

        for cell in column_cells:

            try:

                if cell.value is None:
                    continue

                length = len(str(cell.value))

                if length > max_length:
                    max_length = length

            except Exception:
                pass

        width = max_length + 4

        if width < 12:
            width = 12

        if width > 60:
            width = 60

        ws.column_dimensions[
            column_letter
        ].width = width


def auto_row_height(ws):

    for row in ws.iter_rows(min_row=2):

        max_lines = 1

        for cell in row:

            if cell.value:

                lines = str(cell.value).count("\n") + 1

                length = len(str(cell.value))

                estimate = max(
                    lines,
                    int(length / 60) + 1,
                )

                if estimate > max_lines:
                    max_lines = estimate

        ws.row_dimensions[
            row[0].row
        ].height = max(
            22,
            max_lines * 18,
        )


def format_currency(ws):

    for row in range(2, ws.max_row + 1):

        ws.cell(
            row=row,
            column=10,
        ).number_format = '#,##0.00'

        ws.cell(
            row=row,
            column=11,
        ).number_format = '#,##0.00'


def freeze_header(ws):

    ws.freeze_panes = "A2"


def enable_filter(ws):

    ws.auto_filter.ref = ws.dimensions


def apply_sheet_style(ws):

    freeze_header(ws)

    enable_filter(ws)

    auto_width(ws)

    auto_row_height(ws)

    format_currency(ws)

# ==============================================================================
# MAIN
# ==============================================================================

def main():

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    products = load_json(
        PRODUCT_LIST_FILE,
    )

    metadata = load_json(
        PRODUCT_METADATA_FILE,
    )

    vendor_mapping = load_json(
        VENDOR_MAPPING_FILE,
    )

    vendor_lookup = build_vendor_lookup(
        vendor_mapping,
    )

    product_lookup = build_product_lookup(
        metadata,
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Product List"

    write_header(
        worksheet,
    )

    write_products(
        worksheet,
        products,
        vendor_lookup,
        product_lookup,
    )

    apply_sheet_style(
        worksheet,
    )

    workbook.save(
        OUTPUT_FILE,
    )

    print("=" * 60)
    print("RCBT Product Excel Export")
    print("=" * 60)
    print(f"Products : {len(products)}")
    print(f"Output   : {OUTPUT_FILE}")
    print("=" * 60)


# ==============================================================================
# ENTRY
# ==============================================================================

if __name__ == "__main__":
    main()